"""
views/export_view.py
Exportación a Excel (openpyxl) y PDF (reportlab)
"""
import os
from tkinter import filedialog, messagebox
from datetime import datetime


class ExportView:
    #  EXCEL – openpyxl
    @staticmethod
    def to_excel(rows, columns, default_name="export"):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import (Font, PatternFill, Alignment,
                                          Border, Side, GradientFill)
            from openpyxl.utils import get_column_letter
        except ImportError:
            messagebox.showerror("Error",
                "Instale openpyxl:\n  pip install openpyxl")
            return

        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile=f"{default_name}_{datetime.now():%Y%m%d_%H%M%S}.xlsx")
        if not path:
            return

        wb = Workbook()
        ws = wb.active
        ws.title = default_name.capitalize()

        # ── Título ──
        title_cell = ws.cell(row=1, column=1,
                             value=f"TelecomSys – {default_name.upper()}")
        title_cell.font      = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
        title_cell.fill      = PatternFill("solid", fgColor="1A237E")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells(start_row=1, start_column=1,
                       end_row=1, end_column=len(columns))
        ws.row_dimensions[1].height = 30

        # ── Fecha generación ──
        ws.cell(row=2, column=1,
                value=f"Generado: {datetime.now():%d/%m/%Y %H:%M}")
        ws.cell(row=2, column=1).font = Font(name="Calibri", italic=True,
                                             size=9, color="546E7A")
        ws.merge_cells(start_row=2, start_column=1,
                       end_row=2, end_column=len(columns))

        # ── Encabezados ──
        header_fill   = PatternFill("solid", fgColor="1976D2")
        header_font   = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
        header_align  = Alignment(horizontal="center", vertical="center",
                                  wrap_text=True)
        thin = Side(style="thin", color="B0BEC5")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for ci, col in enumerate(columns, start=1):
            cell = ws.cell(row=3, column=ci, value=col)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = header_align
            cell.border    = border
        ws.row_dimensions[3].height = 22

        # ── Datos ──
        fill_even = PatternFill("solid", fgColor="E8EAF6")
        fill_odd  = PatternFill("solid", fgColor="FFFFFF")
        data_font = Font(name="Calibri", size=10)
        data_align = Alignment(vertical="center")

        for ri, row in enumerate(rows, start=4):
            fill = fill_even if ri % 2 == 0 else fill_odd
            for ci, col in enumerate(columns, start=1):
                val = row.get(col, "")
                if hasattr(val, "isoformat"):
                    val = str(val)[:10]
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.font      = data_font
                cell.fill      = fill
                cell.alignment = data_align
                cell.border    = border

        # ── Autofit columnas ──
        for ci, col in enumerate(columns, start=1):
            max_len = max(
                len(str(col)),
                *[len(str(row.get(col, "") or "")) for row in rows],
                default=10)
            ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 4, 40)

        # ── Congelar primera fila de datos ──
        ws.freeze_panes = "A4"

        wb.save(path)
        messagebox.showinfo("Exportación exitosa",
                            f"Archivo guardado en:\n{path}")

 #  PDF – reportlab
    @staticmethod
    def to_pdf(rows, columns, title="Reporte", fecha_ini=None, fecha_fin=None):
        try:
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib import colors
            from reportlab.lib.units import mm
            from reportlab.platypus import (SimpleDocTemplate, Table,
                                             TableStyle, Paragraph,
                                             Spacer)
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.enums import TA_CENTER, TA_LEFT
        except ImportError:
            messagebox.showerror("Error",
                "Instale reportlab:\n  pip install reportlab")
            return

        path = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("PDF", "*.pdf")],
            initialfile=f"reporte_{datetime.now():%Y%m%d_%H%M%S}.pdf")
        if not path:
            return

        doc = SimpleDocTemplate(path, pagesize=landscape(A4),
                                 leftMargin=15*mm, rightMargin=15*mm,
                                 topMargin=15*mm, bottomMargin=15*mm)

        styles = getSampleStyleSheet()
        PRIMARY = colors.HexColor("#1A237E")
        ACCENT  = colors.HexColor("#1976D2")
        LIGHT   = colors.HexColor("#E8EAF6")

        title_style = ParagraphStyle("title",
            parent=styles["Heading1"],
            fontSize=16, textColor=PRIMARY,
            alignment=TA_CENTER, spaceAfter=4)
        sub_style = ParagraphStyle("sub",
            parent=styles["Normal"],
            fontSize=9, textColor=colors.HexColor("#546E7A"),
            alignment=TA_CENTER, spaceAfter=12)

        story = [
            Paragraph("TelecomSys – Conexión Total S.A.", title_style),
            Paragraph(title, title_style),
        ]

        # Rango de fechas si aplica
        rng_txt = ""
        if fecha_ini and fecha_fin:
            rng_txt = f" | Período: {fecha_ini} – {fecha_fin}"
        story.append(Paragraph(
            f"Generado: {datetime.now():%d/%m/%Y %H:%M}{rng_txt}",
            sub_style))
        story.append(Spacer(1, 6*mm))

        # ── Tabla ──
        header_row = [Paragraph(f"<b>{c}</b>", ParagraphStyle(
            "th", fontSize=8, textColor=colors.white,
            alignment=TA_CENTER)) for c in columns]
        data_rows = []
        for row in rows:
            dr = []
            for col in columns:
                val = row.get(col, "") or ""
                if hasattr(val, "isoformat"):
                    val = str(val)[:10]
                dr.append(Paragraph(str(val),
                    ParagraphStyle("td", fontSize=8,
                                   alignment=TA_LEFT)))
            data_rows.append(dr)

        table_data = [header_row] + data_rows
        col_w = (landscape(A4)[0] - 30*mm) / len(columns)
        table = Table(table_data, colWidths=[col_w]*len(columns),
                      repeatRows=1)

        ts = TableStyle([
            # Header
            ("BACKGROUND",  (0, 0), (-1, 0), ACCENT),
            ("TEXTCOLOR",   (0, 0), (-1, 0), colors.white),
            ("FONTSIZE",    (0, 0), (-1, 0), 8),
            ("ALIGN",       (0, 0), (-1, 0), "CENTER"),
            ("TOPPADDING",  (0, 0), (-1, 0), 6),
            ("BOTTOMPADDING",(0,0), (-1, 0), 6),
            # Data
            ("FONTSIZE",    (0, 1), (-1, -1), 8),
            ("TOPPADDING",  (0, 1), (-1, -1), 4),
            ("BOTTOMPADDING",(0,1), (-1, -1), 4),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1),
             [colors.white, LIGHT]),
            # Grid
            ("GRID",        (0, 0), (-1, -1), 0.5, colors.HexColor("#CFD8DC")),
            ("LINEBELOW",   (0, 0), (-1, 0), 1.5, PRIMARY),
        ])
        table.setStyle(ts)
        story.append(table)

        # Pie de página simple
        story.append(Spacer(1, 8*mm))
        story.append(Paragraph(
            f"Total registros: {len(rows)}",
            ParagraphStyle("footer", fontSize=8,
                           textColor=colors.HexColor("#546E7A"),
                           alignment=TA_LEFT)))

        doc.build(story)
        messagebox.showinfo("Exportación exitosa",
                            f"PDF guardado en:\n{path}")
